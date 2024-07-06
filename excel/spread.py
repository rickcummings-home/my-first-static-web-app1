from openpyxl import load_workbook
#workbook = load_workbook('c:/temp/Clustered.xlsx')
workbook = load_workbook('c:/temp/output.xlsx')
print(workbook.get_sheet_names())
for i in workbook.get_sheet_names():
   worksheet = workbook.get_sheet_by_name(i)
   print(worksheet)
   for column in worksheet.iter_cols() :
      for cell in column:
         print(cell.value)
         cell.value = "rick"
#workbook.save("c:/temp/output.xlsx")
    
